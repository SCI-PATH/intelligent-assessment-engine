"""Canonical Topic ID catalog loaded from the skill-hierarchy Excel sheet.

Runtime prefers the cached ``topics.yaml`` so request handlers do not parse
``.xlsx``. ``scripts/sync_skill_catalog.py`` regenerates that cache after the
spreadsheet changes.
"""

from __future__ import annotations

import re
from functools import lru_cache
from importlib.resources import files
from pathlib import Path

import yaml
from pydantic import BaseModel, Field

from iae.core.curriculum import DEFAULT_GRADE, get_chapter_names
from iae.core.settings import get_settings

SKILL_HIERARCHY_SHEET = "Skill Hierarchy"

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "grade": ("grade", "grade level"),
    "chapter_number": ("chapter", "chapter number", "chapter no", "chapter_no"),
    "chapter_title": ("chapter title", "chapter_title", "chapter name", "chapter_name"),
    "part": ("part",),
    "core_concept": ("core concept", "core_concept", "concept"),
    "topic_id": (
        "topic id (canonical)",
        "topic id",
        "topic_id",
        "canonical topic id",
        "canonical_topic_id",
    ),
    "skill": (
        "curriculum reference",
        "skill",
        "skill name",
        "skill title",
        "skill_name",
    ),
    "domain": ("domain",),
    "concept_code": ("concept code", "concept_code", "code"),
}


class TopicRecord(BaseModel):
    grade: int
    topic_id: str
    chapter_title: str
    skill: str = ""
    chapter_number: int | None = None
    part: str = ""
    core_concept: str = ""
    domain: str = ""
    concept_code: str = ""
    extra: dict[str, str] = Field(default_factory=dict)


def normalize_chapter_name(name: str) -> str:
    """Collapse punctuation so Excel titles match curriculum.yaml names."""
    text = (name or "").lower().replace("-", " ").replace("_", " ")
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return " ".join(text.split())


def _norm_header(value: object) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _header_map(header_row: tuple[object, ...]) -> dict[str, int]:
    found: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        label = _norm_header(cell)
        if not label:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field not in found and label in aliases:
                found[field] = idx
    return found


def _cell(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _parse_grade(raw: str) -> int | None:
    text = raw.strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        digits = re.sub(r"[^0-9]", "", text)
        return int(digits) if digits else None


def parse_skill_workbook(path: str | Path) -> list[TopicRecord]:
    """Read the Skill Hierarchy sheet (column-alias tolerant)."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[SKILL_HIERARCHY_SHEET] if SKILL_HIERARCHY_SHEET in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []

    columns = _header_map(tuple(rows[0]))
    missing = [key for key in ("grade", "chapter_title", "topic_id") if key not in columns]
    if missing:
        raise ValueError(
            f"{path} is missing required columns {missing}. Found headers: {list(rows[0])}"
        )

    known_indexes = set(columns.values())
    topics: list[TopicRecord] = []
    for row in rows[1:]:
        if not row:
            continue
        topic_id = _cell(row, columns.get("topic_id"))
        if not topic_id:
            continue
        grade = _parse_grade(_cell(row, columns.get("grade")))
        if grade is None:
            continue
        extra = {
            _norm_header(rows[0][idx]): _cell(row, idx)
            for idx in range(len(row))
            if idx not in known_indexes and _cell(row, idx)
        }
        chapter_number_raw = _cell(row, columns.get("chapter_number"))
        chapter_number = None
        if chapter_number_raw:
            try:
                chapter_number = int(float(chapter_number_raw))
            except ValueError:
                chapter_number = None
        topics.append(
            TopicRecord(
                grade=grade,
                topic_id=topic_id,
                chapter_title=_cell(row, columns.get("chapter_title")),
                skill=_cell(row, columns.get("skill")),
                chapter_number=chapter_number,
                part=_cell(row, columns.get("part")),
                core_concept=_cell(row, columns.get("core_concept")),
                domain=_cell(row, columns.get("domain")),
                concept_code=_cell(row, columns.get("concept_code")),
                extra=extra,
            )
        )
    return topics


def dump_topics_yaml(topics: list[TopicRecord], dest: Path) -> None:
    payload = {"topics": [t.model_dump() for t in topics]}
    dest.write_text(yaml.safe_dump(payload, sort_keys=False, allow_unicode=True), encoding="utf-8")


def _load_cached_yaml() -> list[TopicRecord]:
    try:
        raw = files("iae.config").joinpath("topics.yaml").read_text(encoding="utf-8")
    except FileNotFoundError:
        return []
    data = yaml.safe_load(raw) or {}
    return [TopicRecord(**entry) for entry in data.get("topics") or []]


@lru_cache(maxsize=1)
def get_topics() -> tuple[TopicRecord, ...]:
    cached = _load_cached_yaml()
    if cached:
        return tuple(cached)
    xlsx = Path(get_settings().skills_xlsx_path)
    if xlsx.exists():
        return tuple(parse_skill_workbook(xlsx))
    return tuple()


def get_topic(topic_id: str) -> TopicRecord | None:
    wanted = (topic_id or "").strip()
    for topic in get_topics():
        if topic.topic_id == wanted:
            return topic
    return None


def topics_for_grade(grade: int = DEFAULT_GRADE) -> list[TopicRecord]:
    return [t for t in get_topics() if t.grade == grade]


def topics_for_chapter(chapter_name: str, grade: int = DEFAULT_GRADE) -> list[TopicRecord]:
    needle = normalize_chapter_name(chapter_name)
    return [
        t
        for t in get_topics()
        if t.grade == grade and normalize_chapter_name(t.chapter_title) == needle
    ]


def match_curriculum_chapters(grade: int = DEFAULT_GRADE) -> tuple[list[str], list[str]]:
    """Return (matched curriculum chapters, unmatched Excel titles) for logging."""
    curriculum = {normalize_chapter_name(name): name for name in get_chapter_names(grade)}
    matched: list[str] = []
    unmatched: list[str] = []
    seen_titles: set[str] = set()
    for topic in topics_for_grade(grade):
        if topic.chapter_title in seen_titles:
            continue
        seen_titles.add(topic.chapter_title)
        key = normalize_chapter_name(topic.chapter_title)
        if key in curriculum:
            matched.append(curriculum[key])
        else:
            unmatched.append(topic.chapter_title)
    return matched, unmatched


def describe_topic(topic: TopicRecord) -> str:
    parts = [topic.skill or topic.core_concept or topic.chapter_title]
    if topic.concept_code:
        parts.append(topic.concept_code)
    if topic.domain:
        parts.append(topic.domain)
    return ". ".join(part for part in parts if part)
